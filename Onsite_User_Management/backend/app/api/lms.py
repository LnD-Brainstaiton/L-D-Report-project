"""LMS API endpoints - ALL DATA FROM LOCAL DATABASE ONLY.

IMPORTANT: These endpoints ONLY read from local database cache.
They NEVER make external API calls. Data is synced via cron job at 12am daily.
"""
from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.orm import Session
import logging
from app.db.base import get_db
from app.core.auth import get_current_admin
from app.models.lms_cache import (
    LMSCourseCache,
    LMSCategoryCache,
    LMSCourseEnrollmentCache,
    LMSUserCourseCache,
    LMSUserCache
)
from app.models.lms_user import LMSUserCourse
from app.models.student import Student
from typing import List, Dict, Any, Optional
from datetime import datetime, date

logger = logging.getLogger(__name__)
router = APIRouter()


@router.get("/courses")
async def get_lms_courses(
    include_enrollment_counts: bool = Query(False, description="Include enrollment count for each course"),
    db: Session = Depends(get_db),
    current_admin: Dict[str, Any] = Depends(get_current_admin)
):
    """
    Get all online courses from LOCAL DATABASE ONLY.
    
    Data is synced from LMS via cron job at 12am daily.
    No external API calls are made.
    """
    try:
        # Get courses from local database cache ONLY
        cached_courses = db.query(LMSCourseCache).all()
        
        if not cached_courses:
            logger.info("No courses in local database. Run sync job to populate.")
            return {"courses": [], "message": "No courses in database. Data syncs daily at 12am."}
        
        # Build result
        result = []
        for course in cached_courses:
            course_data = {
                "id": course.id,
                "fullname": course.fullname or "",
                "startdate": course.startdate,
                "enddate": course.enddate,
                "categoryid": course.categoryid,
                "categoryname": course.categoryname or "Unknown",
                "shortname": course.shortname or "",
                "summary": course.summary or "",
                "visible": course.visible or 1,
                "is_mandatory": course.is_mandatory == 1 if course.is_mandatory is not None else False,
            }
            
            # Get enrollment count from local database if requested
            if include_enrollment_counts:
                # Get from LMSUserCourse table (our local enrollments)
                # Join with Student table to get is_active status
                
                # Total enrollment count
                total_count = db.query(LMSUserCourse).filter(
                    LMSUserCourse.lms_course_id == str(course.id)
                ).count()
                
                # Active employees count (is_active = True)
                active_count = db.query(LMSUserCourse).join(
                    Student, LMSUserCourse.student_id == Student.id
                ).filter(
                    LMSUserCourse.lms_course_id == str(course.id),
                    Student.is_active == True
                ).count()
                
                # Previous employees count (is_active = False)
                previous_count = total_count - active_count
                
                course_data["enrollment_count"] = total_count
                course_data["active_enrollment_count"] = active_count
                course_data["previous_enrollment_count"] = previous_count
            
            result.append(course_data)
        
        return {"courses": result}
        
    except Exception as e:
        logger.error(f"Error fetching courses from local database: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching courses: {str(e)}")


@router.get("/courses/{course_id}/enrollments")
async def get_lms_course_enrollments(course_id: int, db: Session = Depends(get_db)):
    """
    Get all enrolled users for a course from LOCAL DATABASE ONLY.
    
    Data is synced from LMS via cron job at 12am daily.
    No external API calls are made.
    """
    try:
        # Get enrollments from local database - using LMSUserCourse table
        enrollments = db.query(LMSUserCourse).filter(
            LMSUserCourse.lms_course_id == str(course_id)
        ).all()
        
        # Build result with student info
        result = []
        for enrollment in enrollments:
            # Get student info
            student = db.query(Student).filter(Student.id == enrollment.student_id).first()
            
            # Get reporting manager info if available
            reporting_manager_email = None
            if student and student.reporting_manager_employee_id:
                reporting_manager = db.query(Student).filter(
                    Student.employee_id == student.reporting_manager_employee_id
                ).first()
                if reporting_manager:
                    reporting_manager_email = reporting_manager.email
            
            user_data = {
                "id": enrollment.id,
                "student_id": enrollment.student_id,  # Our internal student database ID for fetching details
                "username": enrollment.employee_id,
                "employee_id": enrollment.employee_id,
                "fullname": student.name if student else "",
                "email": student.email if student else "",
                "department": student.department if student else "",
                "sbu_name": student.sbu_name if student else "",  # SBU name from ERP
                "designation": student.designation if student else "",
                "reporting_manager_name": student.reporting_manager_name if student else "",
                "reporting_manager_email": reporting_manager_email,
                "progress": enrollment.progress or 0,
                "completed": enrollment.completed,
                "firstaccess": int(enrollment.start_date.timestamp()) if enrollment.start_date else None,
                "lastaccess": int(enrollment.last_access.timestamp()) if enrollment.last_access else None,
                "is_active": student.is_active if student else True,  # Include is_active status
                "sbu_head_employee_id": student.sbu_head_employee_id if student else None,
                "sbu_head_name": student.sbu_head_name if student else None,
                "reporting_manager_employee_id": student.reporting_manager_employee_id if student else None,
                "bs_joining_date": student.bs_joining_date.isoformat() if student and student.bs_joining_date else None,
                "total_experience": student.total_experience if student else None,
                "career_start_date": student.career_start_date.isoformat() if student and student.career_start_date else None,
                "experience_years": student.experience_years if student else 0,
            }
            result.append(user_data)
        
        return {"enrollments": result}
        
    except Exception as e:
        logger.error(f"Error fetching enrollments from local database: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching enrollments: {str(e)}")


@router.get("/users/{username}/courses")
async def get_lms_user_courses(username: str, db: Session = Depends(get_db)):
    """
    Get all courses for a user from LOCAL DATABASE ONLY.
    
    Data is synced from LMS via cron job at 12am daily.
    No external API calls are made.
    """
    try:
        # Get student by employee_id
        student = db.query(Student).filter(Student.employee_id == username).first()
        
        if not student:
            return {"courses": [], "message": f"User {username} not found in database"}
        
        # Get courses from local database - using LMSUserCourse table
        enrollments = db.query(LMSUserCourse).filter(
            LMSUserCourse.student_id == student.id
        ).all()
        
        # Build result
        result = []
        for enrollment in enrollments:
            course_data = {
                "id": enrollment.lms_course_id,
                "fullname": enrollment.course_name or "",
                "shortname": enrollment.course_shortname or "",
                "startdate": int(enrollment.start_date.timestamp()) if enrollment.start_date else None,
                "enddate": int(enrollment.end_date.timestamp()) if enrollment.end_date else None,
                "progress": enrollment.progress or 0,
                "completed": 1 if enrollment.completed else 0,
                "lastaccess": int(enrollment.last_access.timestamp()) if enrollment.last_access else None,
                "category": enrollment.category_name,
            }
            result.append(course_data)
        
        return {"courses": result}
        
    except Exception as e:
        logger.error(f"Error fetching user courses from local database: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error fetching user courses: {str(e)}")


@router.get("/test-connections")
async def test_api_connections(
    db: Session = Depends(get_db),
    current_admin: Dict[str, Any] = Depends(get_current_admin)
):
    """
    Test connections to both LMS and ERP APIs.
    This is ONLY for testing - actual data sync happens via cron job.
    """
    from app.services.erp_service import ERPService
    from app.services.lms_service import LMSService
    from app.core.config import settings
    
    results = {
        "lms": {},
        "erp": {}
    }
    
    # Test LMS connection
    try:
        if not settings.LMS_TOKEN or not settings.LMS_BASE_URL:
            results["lms"] = {
                "connected": False,
                "configured": False,
                "error": "LMS_TOKEN or LMS_BASE_URL is not configured"
            }
        else:
            try:
                users = await LMSService.fetch_all_users()
                results["lms"] = {
                    "connected": True,
                    "configured": True,
                    "url": settings.LMS_BASE_URL,
                    "users_count": len(users),
                    "message": "Successfully connected to LMS API"
                }
            except Exception as e:
                results["lms"] = {
                    "connected": False,
                    "configured": True,
                    "url": settings.LMS_BASE_URL,
                    "error": str(e)
                }
    except Exception as e:
        results["lms"] = {
            "connected": False,
            "configured": False,
            "error": str(e)
        }
    
    # Test ERP connection
    try:
        erp_result = await ERPService.test_connection()
        results["erp"] = erp_result
    except Exception as e:
        results["erp"] = {
            "connected": False,
            "configured": False,
            "error": str(e)
        }
    
    return results


@router.post("/sync")
async def sync_lms_data(
    db: Session = Depends(get_db),
    current_admin: Dict[str, Any] = Depends(get_current_admin)
):
    """
    Manually trigger LMS data sync.
    
    This is the ONLY place where external LMS API calls are made.
    Normally this runs via cron job at 12am daily.
    """
    from app.services.lms_service import LMSService
    from app.services.lms_cache_service import LMSCacheService
    
    stats = {
        "courses_synced": 0,
        "categories_synced": 0,
        "users_synced": 0,
        "enrollments_synced": 0,
        "errors": []
    }
    
    try:
        # 1. Sync all users
        logger.info("Syncing LMS users...")
        try:
            users = await LMSService.fetch_all_users()
            await LMSCacheService.cache_users(db, users)
            stats["users_synced"] = len(users)
        except Exception as e:
            stats["errors"].append(f"Users sync error: {str(e)}")
        
        # 2. Sync courses and categories
        logger.info("Syncing LMS courses and categories...")
        try:
            courses = await LMSService.fetch_lms_courses()
            category_map = await LMSService.fetch_course_categories()
            await LMSCacheService.cache_courses(db, courses, category_map)
            stats["courses_synced"] = len(courses)
            stats["categories_synced"] = len(category_map)
        except Exception as e:
            stats["errors"].append(f"Courses sync error: {str(e)}")
        
        # 3. Sync enrollments for each course
        logger.info("Syncing LMS enrollments...")
        cached_courses = db.query(LMSCourseCache).all()
        for course in cached_courses:
            try:
                enrolled_users = await LMSService.fetch_course_enrollments(course.id)
                
                # Save enrollments to LMSUserCourse table
                for user in enrolled_users:
                    username = user.get("username", "")
                    if not username or not username.upper().startswith("BS"):
                        continue
                    
                    # Find student
                    student = db.query(Student).filter(Student.employee_id == username).first()
                    if not student:
                        continue
                    
                    # Check if enrollment exists
                    existing = db.query(LMSUserCourse).filter(
                        LMSUserCourse.student_id == student.id,
                        LMSUserCourse.lms_course_id == str(course.id)
                    ).first()
                    
                    # Get is_mandatory from cached course (store as integer: 1 or 0)
                    is_mandatory = 1 if course.is_mandatory == 1 else 0
                    
                    # Use course's timecreated as enrollment_time (when course was created = when it was assigned)
                    # This is the best approximation since enrollment API doesn't provide user-specific enrollment timestamps
                    enrollment_timestamp = None
                    if course.timecreated:
                        enrollment_timestamp = datetime.fromtimestamp(course.timecreated)
                    # Fallback: try to get from user API response if available
                    elif 'timecreated' in user:
                        enrollment_timestamp = datetime.fromtimestamp(user['timecreated']) if user.get('timecreated') else None
                    elif 'timestart' in user:
                        enrollment_timestamp = datetime.fromtimestamp(user['timestart']) if user.get('timestart') else None
                    elif 'enrolments' in user and isinstance(user['enrolments'], list) and len(user['enrolments']) > 0:
                        # Check first enrollment for timecreated
                        first_enrol = user['enrolments'][0]
                        if 'timecreated' in first_enrol:
                            enrollment_timestamp = datetime.fromtimestamp(first_enrol['timecreated']) if first_enrol.get('timecreated') else None
                        elif 'timestart' in first_enrol:
                            enrollment_timestamp = datetime.fromtimestamp(first_enrol['timestart']) if first_enrol.get('timestart') else None
                    
                    if existing:
                        # Update
                        existing.course_name = course.fullname
                        existing.course_shortname = course.shortname
                        existing.category_name = course.categoryname
                        existing.is_mandatory = is_mandatory
                        existing.start_date = datetime.fromtimestamp(course.startdate) if course.startdate else None
                        existing.end_date = datetime.fromtimestamp(course.enddate) if course.enddate else None
                        # Update enrollment_time if we got it from API and it's not already set
                        if enrollment_timestamp and not existing.enrollment_time:
                            existing.enrollment_time = enrollment_timestamp
                    else:
                        # Create
                        new_enrollment = LMSUserCourse(
                            student_id=student.id,
                            employee_id=username,
                            lms_course_id=str(course.id),
                            course_name=course.fullname,
                            course_shortname=course.shortname,
                            category_name=course.categoryname,
                            is_mandatory=is_mandatory,
                            start_date=datetime.fromtimestamp(course.startdate) if course.startdate else None,
                            end_date=datetime.fromtimestamp(course.enddate) if course.enddate else None,
                            enrollment_time=enrollment_timestamp,  # Store enrollment timestamp from LMS if available
                        )
                        db.add(new_enrollment)
                    
                    # Update student has_online_course flag
                    if not student.has_online_course:
                        student.has_online_course = True
                    
                    stats["enrollments_synced"] += 1
                
                db.commit()
            except Exception as e:
                stats["errors"].append(f"Enrollment sync error for course {course.id}: {str(e)}")
                db.rollback()
        
        return {
            "message": "LMS sync completed",
            "stats": stats
        }
        
    except Exception as e:
        logger.error(f"LMS sync error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"LMS sync failed: {str(e)}")


@router.get("/courses/{course_id}/check-mandatory")
async def check_course_mandatory(
    course_id: int,
    db: Session = Depends(get_db),
    current_admin: Dict[str, Any] = Depends(get_current_admin)
):
    """
    Check the is_mandatory status of a specific course.
    """
    try:
        course = db.query(LMSCourseCache).filter(LMSCourseCache.id == course_id).first()
        
        if not course:
            raise HTTPException(status_code=404, detail=f"Course with ID {course_id} not found")
        
        return {
            "course_id": course.id,
            "fullname": course.fullname,
            "shortname": course.shortname,
            "is_mandatory": course.is_mandatory,
            "is_mandatory_bool": course.is_mandatory == 1 if course.is_mandatory is not None else False,
            "is_mandatory_type": str(type(course.is_mandatory)),
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error checking course mandatory status: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error checking course: {str(e)}")


@router.put("/courses/{course_id}/mandatory")
async def update_course_mandatory(
    course_id: int,
    is_mandatory: bool = Query(..., description="Set to true for mandatory, false for optional"),
    db: Session = Depends(get_db),
    current_admin: Dict[str, Any] = Depends(get_current_admin)
):
    """
    Update the is_mandatory status of a specific course.
    Note: This updates the local cache. The value will be overwritten on next sync from LMS.
    """
    try:
        course = db.query(LMSCourseCache).filter(LMSCourseCache.id == course_id).first()
        
        if not course:
            raise HTTPException(status_code=404, detail=f"Course with ID {course_id} not found")
        
        # Update is_mandatory (store as 1 or 0)
        course.is_mandatory = 1 if is_mandatory else 0
        db.commit()
        db.refresh(course)
        
        # Also update all related LMSUserCourse records for this course
        from app.models.lms_user import LMSUserCourse
        db.query(LMSUserCourse).filter(
            LMSUserCourse.lms_course_id == str(course_id)
        ).update({"is_mandatory": course.is_mandatory})
        db.commit()
        
        return {
            "course_id": course.id,
            "fullname": course.fullname,
            "is_mandatory": course.is_mandatory,
            "is_mandatory_bool": course.is_mandatory == 1,
            "message": f"Course mandatory status updated to {'mandatory' if is_mandatory else 'optional'}"
        }
    except HTTPException:
        raise
        raise HTTPException(status_code=500, detail=f"Error updating course: {str(e)}")


@router.get("/courses/{course_id}/report")
def generate_lms_course_report(
    course_id: int, 
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """Generate an Excel report for an online course with enrolled students data.
    
    Supports date range filtering based on enrollment date.
    Includes all students with granular status: Not Started, In Progress, Completed.
    """
    import pandas as pd
    import io
    from fastapi.responses import StreamingResponse
    
    # Get course
    course = db.query(LMSCourseCache).filter(LMSCourseCache.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
    
    # Base query
    query = db.query(LMSUserCourse).filter(
        LMSUserCourse.lms_course_id == str(course_id)
    )
    
    # Apply date filter (using enrollment_time or start_date)
    # Note: enrollment_time is preferred if available
    if start_date:
        # Filter where enrollment_time >= start_date OR (enrollment_time is null AND start_date >= start_date)
        # For simplicity, we'll filter by enrollment_time if available, else start_date
        query = query.filter(
            (LMSUserCourse.enrollment_time >= datetime.combine(start_date, datetime.min.time())) |
            ((LMSUserCourse.enrollment_time == None) & (LMSUserCourse.start_date >= datetime.combine(start_date, datetime.min.time())))
        )
    if end_date:
        query = query.filter(
            (LMSUserCourse.enrollment_time <= datetime.combine(end_date, datetime.max.time())) |
            ((LMSUserCourse.enrollment_time == None) & (LMSUserCourse.start_date <= datetime.combine(end_date, datetime.max.time())))
        )
        
    enrollments = query.all()
    
    # Prepare data for Excel
    report_data = []
    for enrollment in enrollments:
        student = db.query(Student).filter(Student.id == enrollment.student_id).first()
        if not student:
            continue
        
        # Calculate status
        if enrollment.completed:
            status = "Completed"
        elif enrollment.progress and enrollment.progress > 0:
            status = "In Progress"
        else:
            status = "Not Started"
        
        report_data.append({
            'Employee ID': student.employee_id,
            'Name': student.name,
            'Email': student.email,
            'Department': student.department or '',
            'Designation': student.designation or '',
            'Course Name': enrollment.course_name,
            'Category': enrollment.category_name,
            'Status': status,
            'Progress': f"{enrollment.progress}%" if enrollment.progress is not None else "0%",
            'Score': '-',
            'Enrollment Date': enrollment.enrollment_time.strftime('%Y-%m-%d %H:%M:%S') if enrollment.enrollment_time else (enrollment.start_date.strftime('%Y-%m-%d %H:%M:%S') if enrollment.start_date else ''),
            'Completion Date': enrollment.completion_date.strftime('%Y-%m-%d %H:%M:%S') if enrollment.completion_date else '',
            'Last Access': enrollment.last_access.strftime('%Y-%m-%d %H:%M:%S') if enrollment.last_access else '',
            'Is Mandatory': 'Yes' if enrollment.is_mandatory else 'No',
        })
    
    # Create DataFrame
    df = pd.DataFrame(report_data)
    
    # If no enrollments, create empty DataFrame with columns
    if df.empty:
        df = pd.DataFrame(columns=[
            'Employee ID', 'Name', 'Email', 'Department', 'Designation',
            'Course Name', 'Category', 'Status', 'Progress', 'Score',
            'Enrollment Date', 'Completion Date', 'Last Access', 'Is Mandatory'
        ])
    
    # Create Excel file in memory
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Participants', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Participants']
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if not df.empty else 0,
                len(str(col))
            )
            column_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    output.seek(0)
    
    # Generate filename
    safe_course_name = "".join(c for c in course.fullname if c.isalnum() or c in (' ', '-', '_')).strip()
    filename = f"{safe_course_name}_Participants_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/report/overall")
def generate_overall_lms_report(
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """Generate a consolidated Excel report for all online courses with summary stats.
    
    Supports date range filtering based on enrollment date.
    """
    import pandas as pd
    import io
    from fastapi.responses import StreamingResponse
    
    # Get all courses from cache
    courses = db.query(LMSCourseCache).all()
    
    report_data = []
    
    for course in courses:
        # Base query for enrollments
        query = db.query(LMSUserCourse).filter(
            LMSUserCourse.lms_course_id == str(course.id)
        )
        
        # Apply date filter
        if start_date:
            query = query.filter(
                (LMSUserCourse.enrollment_time >= datetime.combine(start_date, datetime.min.time())) |
                ((LMSUserCourse.enrollment_time == None) & (LMSUserCourse.start_date >= datetime.combine(start_date, datetime.min.time())))
            )
        if end_date:
            query = query.filter(
                (LMSUserCourse.enrollment_time <= datetime.combine(end_date, datetime.max.time())) |
                ((LMSUserCourse.enrollment_time == None) & (LMSUserCourse.start_date <= datetime.combine(end_date, datetime.max.time())))
            )
            
        enrollments = query.all()
        
        # Calculate stats
        total_enrolled = len(enrollments)
        completed = sum(1 for e in enrollments if e.completed)
        in_progress = sum(1 for e in enrollments if e.progress and e.progress > 0 and not e.completed)
        not_started = total_enrolled - completed - in_progress
        
        completion_rate = (completed / total_enrolled * 100) if total_enrolled > 0 else 0
        
        report_data.append({
            'Course Name': course.fullname,
            'Short Name': course.shortname,
            'Category': course.categoryname,
            'Start Date': datetime.fromtimestamp(course.startdate).strftime('%Y-%m-%d') if course.startdate else 'N/A',
            'End Date': datetime.fromtimestamp(course.enddate).strftime('%Y-%m-%d') if course.enddate else 'N/A',
            'Total Enrolled': total_enrolled,
            'Completed': completed,
            'In Progress': in_progress,
            'Not Started': not_started,
            'Completion Rate': f"{completion_rate:.1f}%",
            'Is Mandatory': 'Yes' if course.is_mandatory else 'No'
        })
        
    df = pd.DataFrame(report_data)
    
    if df.empty:
        df = pd.DataFrame(columns=[
            'Course Name', 'Short Name', 'Category', 'Start Date', 'End Date',
            'Total Enrolled', 'Completed', 'In Progress', 'Not Started',
            'Completion Rate', 'Is Mandatory'
        ])
        
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Online Courses Summary', index=False)
        
        worksheet = writer.sheets['Online Courses Summary']
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if not df.empty else 0,
                len(str(col))
            )
            column_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
    output.seek(0)
    
    filename = f"Online_Courses_Summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/courses/{course_id}/report/summary")
def generate_lms_course_summary_report(
    course_id: int,
    db: Session = Depends(get_db)
):
    """Generate a summary report for a specific online course."""
    import pandas as pd
    import io
    from fastapi.responses import StreamingResponse
    
    course = db.query(LMSCourseCache).filter(LMSCourseCache.id == course_id).first()
    if not course:
        raise HTTPException(status_code=404, detail="Course not found")
        
    # Calculate stats
    total_enrolled = db.query(LMSUserCourse).filter(
        LMSUserCourse.lms_course_id == str(course_id)
    ).count()
    
    completed = db.query(LMSUserCourse).filter(
        LMSUserCourse.lms_course_id == str(course_id),
        LMSUserCourse.completed == True
    ).count()
    
    in_progress = total_enrolled - completed
    
    completion_rate = (completed / total_enrolled * 100) if total_enrolled > 0 else 0
    
    summary_data = [{
        'Course Name': course.fullname,
        'Short Name': course.shortname,
        'Category': course.categoryname,
        'Start Date': datetime.fromtimestamp(course.startdate).strftime('%Y-%m-%d') if course.startdate else 'N/A',
        'End Date': datetime.fromtimestamp(course.enddate).strftime('%Y-%m-%d') if course.enddate else 'N/A',
        'Total Enrolled': total_enrolled,
        'Completed': completed,
        'In Progress': in_progress,
        'Completion Rate': f"{completion_rate:.1f}%",
        'Is Mandatory': 'Yes' if course.is_mandatory else 'No'
    }]
    
    df = pd.DataFrame(summary_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Auto-adjust column widths
        worksheet = writer.sheets['Summary']
        from openpyxl.utils import get_column_letter
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).map(len).max() if not df.empty else 0,
                len(str(col))
            )
            column_letter = get_column_letter(idx + 1)
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
            
    output.seek(0)
    
    safe_course_name = "".join(c for c in course.fullname if c.isalnum() or c in (' ', '-', '_')).strip()
    filename = f"{safe_course_name}_Summary_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
